import os
import openpyxl

path = "g://aa";
fileList = os.listdir(path)
resultTxt = "g://aa.txt"
for filename in fileList:
    print(filename)
    subPath = os.path.join(path,filename)
    wb = openpyxl.load_workbook(subPath)
    ws = wb.worksheets[0]
    #不读标题，从第2行开始读取数据
    for i in range(2, ws.max_row + 1):
        #print(ws.cell(i, 1).value+","+ws.cell(i, 3).value)
        with open(resultTxt, 'a+') as file_object:
            file_object.write(str(ws.cell(i, 1).value)+","+str(ws.cell(i, 3).value)+"\n")